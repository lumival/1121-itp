import json
import requests
import openpyxl

r = requests.get('https://rickandmortyapi.com/api/character')

data = json.loads(r.text)

wb = openpyxl.load_workbook("C:\\Users\\lv_mana\\Desktop\\1121-itp\\day_4\\lecture.xlsx")
wb.create_sheet("rick and morty")
sheet = wb["rick and morty"]
sheet['A1'] = 'Name'
sheet['B1'] = 'Status'
sheet['C1'] = 'Species'
sheet['D1'] = 'Gender'

for i in range(0, 19):
    print(data['results'][i]['name'])


# print(wb.sheetnames)


# wb = openpyxl.load_workbook("C:\\Users\\lv_mana\\Desktop\\1121-itp\\day_4\\lecture.xlsx")
# print(wb.sheetnames) # ['Sheet1', 'Population...']
# wb.create_sheet('rick and morty')
# ws = wb['rick and morty']
# r = requests.get('https://rickandmortyapi.com/api/character')

# for i in range(20):
#     ws['A' + str(i+1)] = data['results'][i]['name']
#     ws['B' + str(i+1)] = data['results'][i]['Status']
#     ws['C' + str(i+1)] = data['results'][i]['Species']
#     ws['D' + str(i+1)] = data['results'][i]['Gender']